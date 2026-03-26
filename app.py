import streamlit as st
import openpyxl
from io import BytesIO
import zipfile
import os
import re
import tempfile
from openpyxl.utils import get_column_letter

# --- CONTROLE DE VERSÃO ---
VERSAO_SISTEMA = "1.0.2"
DATA_ATUALIZACAO = "Fevereiro/2026"

# --- CONFIGURAÇÃO DA PÁGINA ---
st.set_page_config(page_title="Extrator Análise", layout="centered")

# --- ÁREA PRINCIPAL ---
st.title("Extrator Puro de Medições")
st.caption(f"**Versão {VERSAO_SISTEMA}** ({DATA_ATUALIZACAO}) | Gera planilha limpa, mantém itens separados por etapa e preserva a ordem original das abas do projeto.")

# --- FUNÇÕES AUXILIARES ---
def limpar_codigo(valor):
    if valor is None: return ""
    return str(valor).strip()

def para_float(valor):
    if valor is None: return 0.0
    if isinstance(valor, (int, float)): return float(valor)
    try:
        return float(str(valor).replace(',', '.'))
    except:
        return 0.0

def filtrar_abas_ruas(sheet_names):
    abas_sistema = [
        "DADOS_OBRA", "CSV_GLOBAL", "CSV - CARTILHA", "DADOS_GLOBAL_INICIAL",
        "RESUMO", "CRONOGRAMA", "BDI", "CPU", "ENCARGOS",
        "FAIXAS_ULTIMA_MEDIÇÃO", "CRONOGRAMA_EMPRESA", "GRANDES_ITENS_EMPRESA",
        "DESCRIÇÕES_ETAPAS_EMPRESA", "GRANDES_ITENS", "DESCRIÇÕES_ETAPAS_EDITAL",
        "DMT", "CRONOGRAMA_EDITAL", "ANEXO_V_(ES)", "INFORMATIVO",
        "INSTRUÇÕES", "PLANO_AMOSTRAGEM", "CARTILHA_GLOBAL_PAV",
        "NOVOS_TRAÇOS_CBUQ", "VIAB-PAV", "VIAB-PRAÇA", "ENSAIOS_DE_ORÇAMENTO",
        "COMPOSIÇÕES_COMPLEMENTARES", "SERVIÇOS", "INSUMOS", "DERPR", "DER_MAT", "PREFEITURAS"
    ]
    return [s for s in sheet_names if s.strip().upper() not in abas_sistema]

# --- PROCESSAMENTO (EXTRAÇÃO PURA) ---
def extrair_dados_puros(arquivo_analise):
    # SALVAMENTO TEMPORÁRIO FÍSICO
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(arquivo_analise.getvalue())
        tmp_path = tmp.name

    try:
        wb_analise = openpyxl.load_workbook(tmp_path, data_only=True)
        
        wb_limpo = openpyxl.Workbook()
        
        # 1. Aba Principal de Dados
        ws_limpo = wb_limpo.active
        ws_limpo.title = "Dados_Para_Copiar"
        
        # 0. NOME DO ARQUIVO E LISTA DE RUAS
        abas_ruas = filtrar_abas_ruas(wb_analise.sheetnames)
        nome_municipio = "MUNICIPIO"
        num_sam = "00"
        
        if abas_ruas:
            ws_info = wb_analise[abas_ruas[0]]
            val_h5 = ws_info['H5'].value
            if val_h5: nome_municipio = str(val_h5).strip().replace(" ", "_").upper()
            val_k5 = ws_info['K5'].value
            if val_k5: num_sam = str(val_k5).strip()

        nome_arquivo_final = f"Dados_Limpos_{nome_municipio}_sam{num_sam}.xlsx"

        # 2. Nova Aba: Índice de Ruas
        ws_lista_ruas = wb_limpo.create_sheet(title="Lista_de_Ruas")
        ws_lista_ruas.cell(row=1, column=1).value = "Nº"
        ws_lista_ruas.cell(row=1, column=2).value = "Nome da Aba (Rua/Trecho)"
        ws_lista_ruas.column_dimensions['B'].width = 40 
        
        for idx, nome_rua in enumerate(abas_ruas, start=1):
            ws_lista_ruas.cell(row=idx+1, column=1).value = idx
            ws_lista_ruas.cell(row=idx+1, column=2).value = nome_rua

        # A. LISTA MESTRA
        aba_origem_csv = None
        if 'CSV_GLOBAL' in wb_analise.sheetnames: aba_origem_csv = 'CSV_GLOBAL'
        elif 'CSV - Cartilha' in wb_analise.sheetnames: aba_origem_csv = 'CSV - Cartilha'
        
        mapa_linha_exata = {} 
        mapa_codigo_lista = {} 

        if aba_origem_csv:
            ws_orig = wb_analise[aba_origem_csv]
            linha_dest = 10
            seq_contador = 1
            
            ws_limpo.cell(row=9, column=6).value = "Item"
            ws_limpo.cell(row=9, column=7).value = "Código"
            ws_limpo.cell(row=9, column=8).value = "Descrição"
            
            for r_orig in range(11, ws_orig.max_row + 1):
                filtro = ws_orig.cell(row=r_orig, column=4).value 
                
                if filtro and str(filtro).strip().upper() == 'X':
                    raw_code = ws_orig.cell(row=r_orig, column=6).value
                    code_key = limpar_codigo(raw_code)
                    
                    if code_key:
                        mapa_linha_exata[r_orig] = {"linha_dest": linha_dest, "codigo": code_key}
                        mapa_codigo_lista.setdefault(code_key, []).append(linha_dest)

                    for c_orig in range(5, 12): 
                        c_dest = c_orig + 1 
                        if c_orig == 5: 
                            v_orig = ws_orig.cell(row=r_orig, column=c_orig).value
                            if v_orig and (isinstance(v_orig, (int, float)) or str(v_orig).strip().isdigit()):
                                ws_limpo.cell(row=linha_dest, column=c_dest).value = seq_contador
                                seq_contador += 1
                            else:
                                ws_limpo.cell(row=linha_dest, column=c_dest).value = v_orig
                        else:
                            val = ws_orig.cell(row=r_orig, column=c_orig).value
                            if val is not None:
                                ws_limpo.cell(row=linha_dest, column=c_dest).value = val
                    linha_dest += 1

        # B. QUANTIDADES POR RUA
        coluna_atual = 19 # S
        
        for nome_aba in abas_ruas:
            ws_rua = wb_analise[nome_aba]
            
            ws_limpo.cell(row=8, column=coluna_atual).value = nome_aba
            
            for linhas_dest in mapa_codigo_lista.values():
                for linha_item in linhas_dest:
                    ws_limpo.cell(row=linha_item, column=coluna_atual).value = 0.0

            ocorrencias_vistas = {} 

            max_r_rua = min(ws_rua.max_row, 5000)
            
            for r in range(11, max_r_rua + 1):
                filtro_rua = ws_rua.cell(row=r, column=3).value
                if filtro_rua and str(filtro_rua).strip().upper() == 'X':
                    raw_code_rua = ws_rua.cell(row=r, column=7).value
                    key_rua = limpar_codigo(raw_code_rua)
                    
                    qtd_rua = para_float(ws_rua.cell(row=r, column=20).value)
                    
                    if key_rua:
                        linha_alvo = None
                        
                        if r in mapa_linha_exata and mapa_linha_exata[r]["codigo"] == key_rua:
                            linha_alvo = mapa_linha_exata[r]["linha_dest"]
                        else:
                            idx = ocorrencias_vistas.get(key_rua, 0)
                            lista_linhas = mapa_codigo_lista.get(key_rua, [])
                            
                            if idx < len(lista_linhas):
                                linha_alvo = lista_linhas[idx]
                            elif len(lista_linhas) > 0:
                                linha_alvo = lista_linhas[-1]

                        if linha_alvo is not None:
                            valor_atual = ws_limpo.cell(row=linha_alvo, column=coluna_atual).value
                            if valor_atual is None: valor_atual = 0.0
                            ws_limpo.cell(row=linha_alvo, column=coluna_atual).value = valor_atual + qtd_rua
                            
                        ocorrencias_vistas[key_rua] = ocorrencias_vistas.get(key_rua, 0) + 1

            coluna_atual += 1
            # O limite de colunas foi removido daqui!

        # C. FÓRMULAS DE CHECAGEM
        ultima_coluna_ruas_idx = coluna_atual - 1
        if ultima_coluna_ruas_idx >= 19:
            letra_ultima = get_column_letter(ultima_coluna_ruas_idx)
            for linhas_dest in mapa_codigo_lista.values():
                for linha_item in linhas_dest:
                    formula = f'=IF(K{linha_item}=0,"-",IF(ROUND(K{linha_item},2)=ROUND(SUM(S{linha_item}:{letra_ultima}{linha_item}),2),"Ok","Verificar"))'
                    ws_limpo.cell(row=linha_item, column=18).value = formula

        dados_finais = BytesIO()
        wb_limpo.save(dados_finais)
        wb_limpo.close()
        wb_analise.close()
        
        return dados_finais, nome_arquivo_final

    finally:
        if os.path.exists(tmp_path):
            try:
                os.unlink(tmp_path)
            except:
                pass

# --- INTERFACE ---
st.subheader("1. Arquivo(s) de Análise da Prefeitura")
arquivos_analise = st.file_uploader("Arquivos .xlsx", accept_multiple_files=True)

st.markdown("---")
action_container = st.empty()

if arquivos_analise:
    if action_container.button("Extrair Dados Limpos"):
        action_container.button("Processando...", disabled=True)
        
        with st.spinner("Gerando planilhas de extração (pode demorar em arquivos grandes)..."):
            try:
                total_arquivos = len(arquivos_analise)
                
                if total_arquivos == 1:
                    arq = arquivos_analise[0]
                    res, nome_gerado = extrair_dados_puros(arq)
                    
                    st.success(f"Sucesso! Arquivo extraído: {nome_gerado}")
                    
                    st.download_button(
                        label="Baixar Dados Limpos",
                        data=res.getvalue(),
                        file_name=nome_gerado,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                else:
                    zip_buffer = BytesIO()
                    with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
                        for arq in arquivos_analise:
                            res, nome_gerado = extrair_dados_puros(arq)
                            zip_file.writestr(nome_gerado, res.getvalue())
                    
                    st.success(f"{total_arquivos} extrações geradas com sucesso!")
                    
                    st.download_button(
                        label="Baixar Lote ZIP",
                        data=zip_buffer.getvalue(),
                        file_name="Lote_Dados_Limpos.zip",
                        mime="application/zip"
                    )
                    
            except Exception as e:
                action_container.empty()
                st.error(f"Ocorreu um erro: {e}")
                import traceback
                st.error(traceback.format_exc())
                if st.button("Tentar Novamente"):
                     st.rerun()
else:
    st.button("Extrair Dados Limpos", disabled=True, help="Faça upload dos arquivos primeiro")